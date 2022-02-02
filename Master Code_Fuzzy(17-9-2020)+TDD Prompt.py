import modin.pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog
from tkinter import messagebox as mb
import os
import sys
import datetime
import xlsxwriter
from openpyxl import load_workbook
import numpy as np

from fuzzywuzzy import fuzz, process

#SUMMARY

#_____________________________________________________________________________________________________
# This script is run for mycom mobinet DPR and CBS reports of a given circle and returns a sorted
# excel sheet by ECGI ID numbers for each cell, appended column-wise.

# Current implementation involves getting the unique ECGI IDs for the first 100 string sorted ECGIs and
# populating a site handbook for each individual cell
#_____________________________________________________________________________________________________



import time

def TicTocGenerator():
    # Generator that returns time differences
    ti = 0           # initial time
    tf = time.time() # final time
    while True:
        ti = tf
        tf = time.time()
        yield tf-ti # returns the time difference

TicToc = TicTocGenerator() # create an instance of the TicTocGen generator

# This will be the main function through which we define both tic() and toc()
def toc(tempBool=True):
    # Prints the time difference yielded by generator instance TicToc
    tempTimeInterval = next(TicToc)
    if tempBool:
        print( "Elapsed time: %f seconds.\n" %tempTimeInterval )

def tic():
    # Records a time in TicToc, marks the beginning of a time interval
    toc(False)

#_______________________________________________________________________________________________________


NODECOUNTER = 0
gis = pd.DataFrame()
mycom = pd.DataFrame()
dpr = pd.DataFrame()
cbs = pd.DataFrame()
mobinet = pd.DataFrame()


FinalDF = pd.DataFrame()

x = datetime.datetime.now()

filename = "CellsAll.xlsx"

filedets = pd.DataFrame({'Path': [
    r"C:\Users\b0219084\PycharmProjects\Radio Planning Project(14-Jul-2020)\Presentation\Match Report_" + filename],
                         'Date': [x.strftime("%d-%m-%Y")],
                         'Time': [x.strftime("%H:%M")]})
workbook = xlsxwriter.Workbook(filename)
workbook = workbook.add_worksheet()
writer = pd.ExcelWriter(workbook, engine='xlsxwriter')



report={
    "1": "GIS",
    "2": "MyCom",
    "3": "DPR",
    "4": "CBS_FDD",
    "5": "CBS_TDD",
    "6": "Mobinet 4G"
}


column_dict={
    "1":["Site Address", "eNB Model No","RRU Model No","eNB Software version", "Band", "SRAN", "Latitude", "Longitude","Diplexer (Yes/No)", "Diplexer Purpose  (Antenna sharing : 2G & 4G )", "Antenna Shared with 2G/3G (Yes/No)", "Antenna  Make", "Antenna  Model", "Antenna Height (m)", "RET (Yes/No)", "Tower Type (GBT/RTT/RTP/Wall Mounted/GBP/COW/ NOW/GBT+Revamp/RTT+Revamp/IBS)", "Tower Height (m)", "Building Height (m)", "Azimuth", "Site Principal Owner", "ICR category"],
    "2":["4G Data Volume [GB]", "4G Data Volume [GB] [CDBH]","DL User Throughput_Kbps [CDBH]", "Average number of used DL PRBs [CDBH]", "Avg Connected User [CDBH]", "E-UTRAN Average CQI [CDBH]", "VoLTE Traffic", "VoLTE Traffic [CBBH]", "No of Hours having Data Volume >80% of CDBH Data Volume", "No of Hours with PRB Utilization DL >80%", "No of Hours with PRB Utilization DL >80% and User Throughput DL < 2Mbps", "Avg Connected User", "Average number of used DL PRBs", "DL User Throughput_Kbps", "Data Volume UL - Total", "Data Volume UL - Total [CDBH]",	"E-UTRAN Average CQI", "Average number of used DL PRBs [CUBH]",	"Avg Connected User [CUBH]", "DL User Throughput_Kbps [CUBH]",	"E-UTRAN Average CQI [CUBH]" ],
    "4":['eNode-B ID', 'SITENAME', 'eNode-B Cell ID', 'ECGI', 'Towns', 'T50/SC/NSC', 'DOI-Date of Integration (DD-MM-YY)', 'MS 1.0', 'MS 2.0', 'Site TYPE (IM-Indoor Macro,OM-Outdoor Macro,TT-Tower Top Macro,MI-Micro, IB-IBS)/Small cell', 'eNode-B ID Type- Main Cabinet', 'eNodeB ID TypeSecond Cabinet', 'Power License (in Watt)', 'Existing/New', 'Active /locked', 'Date Since Locked (DD-MM-YY', 'Site Dismantled and Reached at WH (DD-MM-YY)', 'Colocated with 2G (Yes/No)', 'Colocated with 3G (Yes/No)', 'Site status FDD/TDD', 'Lean / NON LeanSITE', 'TECH ID', 'Existing 2G Site ID', 'Existing 3G Site ID', 'Existing 4G TDD Site ID'],
    "5":["Cell Technology", "Mimo Configuration","Downlink Bandwidths", "Uplink Bandwidths","CGI", "Site ID", "SRAN", "Power", "RU", "RU Model", "RU serial number", "Serial Number", "GPS Latitude", "GPS Longitude", "Status"]
    }
#________________________________________________

# File reading functions

def DataRead(Number):

    mb.showinfo("Select File", "Report Category: {}".format(report[str(Number)]))
    FileSelectFlag  = False
    if Number!=5:
        try:
            path= filedialog.askopenfilename()
            print("Selected path:", path, "\n")
            FileSelectFlag = True

        except:
            mb.showinfo("File Selection Error",  "No file detected. Rerun script")
            print("No file was selected")
            sys.exit()
            # Split filename and file extension into two separate variables
        filename, file_extension = os.path.splitext(path)
        file = os.path.basename(path)
        # sheet = pd.read_csv(filename, encoding='cp1252', low_memory=False)
        if file_extension == '.csv':
            sheet = pd.read_csv(file,  encoding='cp1252', low_memory=False)
        elif file_extension == '.xlsx':
            sheet = pd.read_excel(file)

        print("Imported Sheet:-", os.path.basename(path))
        print(sheet.head(5))
        return sheet, FileSelectFlag

    else:
        
            path= filedialog.askopenfilename()
            if len(path) == 0:
                sheet = pd.DataFrame()
                return sheet, FileSelectFlag
            else:              
                print("Selected path:", path, "\n")
                FileSelectFlag = True
                filename, file_extension = os.path.splitext(path)
                file = os.path.basename(path)
                # sheet = pd.read_csv(filename, encoding='cp1252', low_memory=False)
                if file_extension == '.csv':
                    sheet = pd.read_csv(file,  encoding='cp1252', low_memory=False)
                elif file_extension == '.xlsx':
                    sheet = pd.read_excel(file)

                print("Imported Sheet:-", os.path.basename(path))
                print(sheet.head(5))
                return sheet, FileSelectFlag



    
# ____________________________________________
# FIle appending and data formatting done here


def cellmapper(df1, df2, df3, df4, df5):
    global NODECOUNTER, FinalDF
    gisnew = pd.DataFrame()
    cbsnew = pd.DataFrame()
    mobinetnew = pd.DataFrame()
    mcomnew = pd.DataFrame()

    
    gis_cols=get_df_col(1,df1)
    mycom_cols=get_df_col(2,df2)
    cbs_cols=get_df_col(4,df4)
    mobinet_cols=get_df_col(5,mobinet)
    

    
    for i in df4.iloc[:,:]["ECGI_1"].unique():
        print("Searching Database")
        tic()
        gisnew = df1[df1["ECGI_1"] == i]
        mcomnew = df2[df2["ECGI_1"] == i]
        # dprnew = df3[df3["CGI ID"] == i]
        cbsnew = df4[df4["ECGI_1"] == i]
        mobinetnew = df5[df5["ECGI_1"] == i]
        # giscols  = df1.columns[[2, 26, 27,47,16,17,18,20, 24,31, 32, 34, 35, 36, 37,41, 44, 45,46, 50, 54]]
        # mcomcols = df2.columns[a+1:a+21]
        # cbscols  = df4.columns[[2,3,7,9,10,12,13,14,15,16,17,18,19,20,21,26,27,29,30,33,34,35,36,37]]
        # mobinetcols = df5.columns[[6, 11,12,13,15,18,22,23,24,25,26,27,32]]
        #toc()
        
        DF = pd.DataFrame()
        DF2= pd.DataFrame()
        

        #print("Merging Gis and cbs")
        #tic()
        gisnew  = gisnew.iloc[:, gis_cols]
        gisnew = gisnew.reset_index()
        
        ECGI=cbsnew['ECGI_1']
        ECGI= ECGI.reset_index()

        DF = pd.concat([ECGI, gisnew], axis=1)
        
        cbsnew  = cbsnew.iloc[:,cbs_cols]
        cbsnew = cbsnew.reset_index()
        DF= pd.concat([DF, cbsnew], axis =1)
        #toc()
        
        #print("Mergin Mcom and Mobinet")
        #tic()
        mcomnew = mcomnew.iloc[:, mycom_cols]
        mcomnew = mcomnew.reset_index()
        
        mobinetnew  = mobinetnew.iloc[:,mobinet_cols]
        mobinetnew = mobinetnew.reset_index()
        
        DF2= pd.concat([mcomnew,mobinetnew], axis =1)
        #toc()
        
        #print("Merge All and append to Final DF")
        #tic()
        DF= pd.concat([DF,DF2], axis=1)
        del DF2
        DF= DF.reset_index()
        del DF["index"]
        del DF["level_0"]
        
        FinalDF = FinalDF.append(DF)
        NODECOUNTER =NODECOUNTER+1
        print('NODECOUNTER= ', NODECOUNTER)
        toc()
        
    
#_______________________HELPER FUNCTIONS___________________________

#_____________________________________________________________________________________________
#_________________RENAME ALL COLUMNS WITH CGI (CGI/LTE CGI/MV-ECGI => ECGI_1__________________

def column_rename(df):
    dfcols = []
    count = 1
    for columnnum in range(len(df.columns)):
        if 'cgi' in str(df.columns[columnnum]).lower():  
            dfcols.append(f'ECGI_{count}')
            count+=1
            continue
        dfcols.append(df.columns[columnnum])  
    df.columns = dfcols
    return df
#_____________________________________________________________________________________________
#_________REMOVE EXTRA ECGI ZEROS  (404-41-0093234-16 => 404-41-93234-16)_____________________

def stripFunc(df):

    for ecgi in df['ECGI_1']:
        try:
            lastIndex  = ecgi.rindex('-')
            startIndex = ecgi.rindex('-', 0, lastIndex)
            substr = ecgi[startIndex+1:lastIndex]
            index = 0
            for number in range(len(substr)):
                if substr[number] == '0':
                    continue
                else:
                    index = number
                    break
            if index == 0:
                continue
            else:
                substr2 = substr[index:len(substr)]
                ecgi2 = ecgi.replace(substr, substr2)
                df['ECGI_1'].replace(ecgi, ecgi2)
                print('{}=>{}'.format(ecgi,ecgi2))
        except(LookupError, ValueError, IndexError, AttributeError):
            continue

       
    return df
       
def stripFunc_improved(df):

    for ecgi in df['ECGI_1']:
        try:
            strList = ecgi.split('-')

            for elementNum in range(len(strList)):
                strList[elementNum] = strList[elementNum].lstrip('0')
            ecgi2 ='-'.join(strList)

            if ecgi == ecgi2:
                continue
            else:
                df['ECGI_1'] = df['ECGI_1'].replace(ecgi, ecgi2)
                print('{}=>{}'.format(ecgi,ecgi2))
        except(LookupError, ValueError, IndexError, AttributeError):
            continue

       
    return df
#_____________________________________________________________________________________________
#________GET COLUMN INDICES USING FUZZY____________________
  
#Returns a list of indices which are in the query list and search list
def get_df_col(df_num,df):
    
    search_list=df.columns
    query_list=column_dict[str(df_num)]
    indexList=[]
    for entry in query_list:
        temp_tuple=process.extractOne(entry,search_list)
        col_num=df.columns.get_loc(str(temp_tuple[0]))
        indexList.append(col_num)
        
    # for i in range(len(indexList)-1):  
    #     if indexList[i+1]==indexList[i]:
    #         indexList[i+1]=indexList[i]+1 
    return indexList
  
#_____________________________________________________________________________________________
#________ APPEND INFORMATION TO AN EXISTING EXCEL SHEET (UNUSED)____________________
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow ,index=None, **to_excel_kwargs)

    # save the workbook
    writer.save()  
#_____________________________________________________________________________________________



def main():
        global gis, cbs, dpr, mycom, mobinet, FinalDF, reports,cbs_fdd

        root = tk.Tk()
        root.withdraw()  # Closes default window of tkinter package
# *** Gis ***
        isCBSTDD=mb.askyesno("Question", "Do we have CBS TDD?")
        [gis, _] = DataRead(1)

# *** MyCom ***
        [mycom, _] = DataRead(2)
        
# *** CBS ***
        [cbs_fdd, Flag1] = DataRead(4)
        
        Flag2=False
        if isCBSTDD:
            
            [cbs_tdd, Flag2] = DataRead(5)
       
        # if Flag1 and Flag2 == True:
        #     cbs = pd.concat([cbs_fdd, cbs_tdd], axis = 0)
        # else:
        #     if Flag1 ==True:
        #         cbs = cbs_fdd
        #         del cbs_fdd, cbs_tdd
        #     if Flag2 ==True:
        #         cbs = cbs_tdd
        #         del cbs_fdd, cbs_tdd
                

#_____________________________________________________________
#__BLOCK RESERVED FOR MERGED HEADERS IN CBS_FDD: WILL CHANGE______
        # cbs_fdd.columns = cbs_fdd.iloc[0]
        # cbs_fdd = cbs_fdd.drop(0)
#_____________________________________________________________

#_____________________________________________________________
#__BLOCK RESERVED FOR MERGED HEADERS IN CBS_TDD: WILL CHANGE______
        # if isCBSTDD:
            
        #     cbs_tdd.columns = cbs_tdd.iloc[0]
        #     cbs_tdd = cbs_tdd.drop(0)
#_____________________________________________________________




# *** Mobinet ***
        [mobinet,_] = DataRead(6)


        #
        # report = {
        #     "1": "GIS",
        #     "2": "MyCom",
        #     "3": "DPR",
        #     "4": "CBS",
        #     "5": "Mobinet 4G"
        # }
        #
# *** HELPER FUNCTION CALLS ***  
        if isCBSTDD:
            
            cbs_tdd = column_rename(cbs_tdd)
            cbs_tdd = stripFunc_improved(cbs_tdd)

        cbs_fdd = column_rename(cbs_fdd)
        cbs_fdd = stripFunc_improved(cbs_fdd)

        gis = column_rename(gis)
        gis = stripFunc_improved(gis)

        mycom = column_rename(mycom)
        mycom = stripFunc_improved(mycom)

        mobinet = column_rename(mobinet)
        mobinet = stripFunc_improved(mobinet)

      
#SORTING VALUES ACCORDING TO ECGI VALUE IN ASCENDING ORDER 
        gis.sort_values(by=["ECGI_1"], ascending=True)
        mycom.sort_values(by=["ECGI_1"], ascending=True)
        cbs_fdd.sort_values(by=["ECGI_1"], ascending=True)
        if isCBSTDD:
            
            cbs_tdd.sort_values(by=["ECGI_1"], ascending=True)
        mobinet.sort_values(by=["ECGI_1"], ascending=True)

# *** CELLMAPPER FUNCTION ***        
        if Flag1:
            cellmapper(gis, mycom, dpr, cbs_fdd, mobinet)
            FinalDF = FinalDF.set_index("ECGI_1")
            FinalDF.to_csv("CellsAll_FDD_{}.csv".format(x.strftime("%d-%m-%Y")))

        FinalDF = pd.DataFrame()

        if Flag2:
            cellmapper(gis, mycom, dpr, cbs_tdd, mobinet)
            FinalDF = FinalDF.set_index("ECGI_1")
            FinalDF.to_csv("CellsAll_TDD_{}.csv".format(x.strftime("%d-%m-%Y")))
 

if __name__ == "__main__":
    main()

